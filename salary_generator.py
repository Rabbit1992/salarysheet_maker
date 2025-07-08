import streamlit as st
import pandas as pd
import io
from datetime import datetime, date
import os
from openpyxl import load_workbook
from copy import copy
import calendar

def get_chinese_holidays_2024():
    """获取2024年中国法定节假日列表"""
    holidays = [
        # 元旦
        date(2024, 1, 1),
        # 春节
        date(2024, 2, 10), date(2024, 2, 11), date(2024, 2, 12), 
        date(2024, 2, 13), date(2024, 2, 14), date(2024, 2, 15), date(2024, 2, 16), date(2024, 2, 17),
        # 清明节
        date(2024, 4, 4), date(2024, 4, 5), date(2024, 4, 6),
        # 劳动节
        date(2024, 5, 1), date(2024, 5, 2), date(2024, 5, 3), date(2024, 5, 4), date(2024, 5, 5),
        # 端午节
        date(2024, 6, 10),
        # 中秋节
        date(2024, 9, 15), date(2024, 9, 16), date(2024, 9, 17),
        # 国庆节
        date(2024, 10, 1), date(2024, 10, 2), date(2024, 10, 3), 
        date(2024, 10, 4), date(2024, 10, 5), date(2024, 10, 6), date(2024, 10, 7)
    ]
    return holidays

def get_chinese_holidays_2025():
    """获取2025年中国法定节假日列表"""
    holidays = [
        # 元旦
        date(2025, 1, 1),
        # 春节
        date(2025, 1, 28), date(2025, 1, 29), date(2025, 1, 30), 
        date(2025, 1, 31), date(2025, 2, 1), date(2025, 2, 2), date(2025, 2, 3),
        # 清明节
        date(2025, 4, 5), date(2025, 4, 6), date(2025, 4, 7),
        # 劳动节
        date(2025, 5, 1), date(2025, 5, 2), date(2025, 5, 3), date(2025, 5, 4), date(2025, 5, 5),
        # 端午节
        date(2025, 5, 31), date(2025, 6, 1), date(2025, 6, 2),
        # 中秋节
        date(2025, 10, 6),
        # 国庆节
        date(2025, 10, 1), date(2025, 10, 2), date(2025, 10, 3), 
        date(2025, 10, 4), date(2025, 10, 5), date(2025, 10, 7), date(2025, 10, 8)
    ]
    return holidays

def is_holiday_or_weekend(date_obj):
    """判断日期是否为法定节假日或周末"""
    if not isinstance(date_obj, date):
        return False, "工作日"
    
    # 获取对应年份的节假日
    if date_obj.year == 2024:
        holidays = get_chinese_holidays_2024()
    elif date_obj.year == 2025:
        holidays = get_chinese_holidays_2025()
    else:
        holidays = []
    
    # 判断是否为法定节假日
    if date_obj in holidays:
        return True, "法定节假日"
    
    # 判断是否为周末
    if date_obj.weekday() >= 5:  # 5=周六, 6=周日
        return True, "休息日"
    
    return False, "工作日"

def parse_date_from_string(date_str):
    """从字符串中解析日期"""
    if pd.isna(date_str) or not date_str:
        return None
    
    date_str = str(date_str).strip()
    
    # 尝试多种日期格式
    date_formats = [
        '%Y-%m-%d',
        '%Y/%m/%d',
        '%Y年%m月%d日',
        '%m/%d/%Y',
        '%d/%m/%Y',
        '%Y-%m-%d %H:%M:%S',
        '%Y/%m/%d %H:%M:%S'
    ]
    
    for fmt in date_formats:
        try:
            parsed_date = datetime.strptime(date_str, fmt).date()
            return parsed_date
        except ValueError:
            continue
    
    # 如果是pandas的Timestamp对象
    try:
        if hasattr(date_str, 'date'):
            return date_str.date()
        elif isinstance(date_str, datetime):
            return date_str.date()
    except:
        pass
    
    return None

def load_salary_template():
    """加载工资表模板"""
    try:
        template_path = "工资表模板.xlsx"
        if os.path.exists(template_path):
            # 工资表模板第五行为标题，数据从第六行开始，所以使用header=4
            df = pd.read_excel(template_path, header=4)
            
            # 过滤掉空行和无用列
            df = df.dropna(subset=['姓名'])
            
            # 清理列名，移除无用的Unnamed列
            df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
            
            st.success(f"成功加载工资表模板，找到 {len(df)} 名员工")
            return df, template_path
        else:
            st.error(f"找不到工资表模板文件: {template_path}")
            return None, None
    except Exception as e:
        st.error(f"加载工资表模板时出错: {str(e)}")
        return None, None

def load_leave_data(uploaded_file):
    """加载休假数据"""
    if uploaded_file is not None:
        try:
            # 尝试不同的header位置来找到正确的数据行
            for header_row in [0, 1, 2, 3, 4]:
                try:
                    df = pd.read_excel(uploaded_file, header=header_row)
                    # 检查是否包含必要的列
                    if '创建人' in df.columns and ('请假类型' in df.columns or '时长' in df.columns):
                        # 过滤掉空行
                        df = df.dropna(subset=['创建人'])
                        st.info(f"成功读取休假数据，找到 {len(df)} 条记录")
                        return df
                except:
                    continue
            
            # 如果没有找到合适的格式，尝试手动解析
            df = pd.read_excel(uploaded_file, header=None)
            # 查找包含'创建人'的行
            for i, row in df.iterrows():
                if '创建人' in row.values:
                    # 使用这一行作为列名
                    df.columns = df.iloc[i]
                    df = df.iloc[i+1:].reset_index(drop=True)
                    df = df.dropna(subset=['创建人'])
                    st.info(f"成功解析休假数据，找到 {len(df)} 条记录")
                    return df
            
            st.error("无法在休假表中找到'创建人'列，请检查文件格式")
            return None
        except Exception as e:
            st.error(f"读取休假数据时出错: {str(e)}")
            return None
    return None

def load_overtime_data(uploaded_file):
    """加载加班数据"""
    if uploaded_file is not None:
        try:
            # 尝试不同的header位置来找到正确的数据行
            for header_row in [0, 1, 2, 3, 4]:
                try:
                    df = pd.read_excel(uploaded_file, header=header_row)
                    # 检查是否包含必要的列
                    if '创建人' in df.columns and '时长' in df.columns:
                        # 过滤掉空行
                        df = df.dropna(subset=['创建人'])
                        st.info(f"成功读取加班数据，找到 {len(df)} 条记录")
                        return df
                except:
                    continue
            
            # 如果没有找到合适的格式，尝试手动解析
            df = pd.read_excel(uploaded_file, header=None)
            # 查找包含'创建人'的行
            for i, row in df.iterrows():
                if '创建人' in row.values:
                    # 使用这一行作为列名
                    df.columns = df.iloc[i]
                    df = df.iloc[i+1:].reset_index(drop=True)
                    df = df.dropna(subset=['创建人'])
                    st.info(f"成功解析加班数据，找到 {len(df)} 条记录")
                    return df
            
            st.error("无法在加班表中找到'创建人'列，请检查文件格式")
            return None
        except Exception as e:
            st.error(f"读取加班数据时出错: {str(e)}")
            return None
    return None

def process_leave_data(result_df, leave_data):
    """处理休假数据并更新到工资表现有列中"""
    if leave_data is not None:
        # 检查必要的列是否存在
        required_leave_columns = ['创建人', '请假类型', '时长']
        missing_columns = [col for col in required_leave_columns if col not in leave_data.columns]
        
        if missing_columns:
            st.error(f"休假数据文件缺少必要的列: {', '.join(missing_columns)}")
            st.error(f"当前文件包含的列: {', '.join(leave_data.columns.tolist())}")
            st.error("请确保休假数据文件包含以下列：创建人、请假类型、时长")
            return result_df
        
        # 不再过滤审批结果，处理所有休假数据
        st.info(f"将处理所有 {len(leave_data)} 条休假记录（不考虑审批状态）")
        
        # 处理时长数据，统一转换为天数
        def parse_duration(duration_str):
            if pd.isna(duration_str):
                return 0
            duration_str = str(duration_str).strip()
            if '天' in duration_str:
                return float(duration_str.replace('天', ''))
            elif '小时' in duration_str or 'h' in duration_str.lower():
                hours = float(duration_str.replace('小时', '').replace('h', '').replace('H', ''))
                return hours / 8  # 按8小时工作日计算
            else:
                try:
                    return float(duration_str)
                except:
                    return 0
        
        leave_data['休假天数'] = leave_data['时长'].apply(parse_duration)
        
        # 为每个员工收集详细的休假记录
        for index, row in result_df.iterrows():
            employee_name = row['姓名']
            employee_leaves = leave_data[leave_data['创建人'] == employee_name]
            
            if not employee_leaves.empty:
                leave_details = []
                total_days = 0
                has_unpaid_leave = False
                
                # 遍历该员工的所有休假记录
                for _, leave_record in employee_leaves.iterrows():
                    leave_type = str(leave_record['请假类型']) if pd.notna(leave_record['请假类型']) else '未知类型'
                    start_time = str(leave_record['开始时间']) if pd.notna(leave_record['开始时间']) and '开始时间' in leave_record else ''
                    end_time = str(leave_record['结束时间']) if pd.notna(leave_record['结束时间']) and '结束时间' in leave_record else ''
                    duration = str(leave_record['时长']) if pd.notna(leave_record['时长']) else ''
                    approval_status = str(leave_record['审批结果']) if pd.notna(leave_record['审批结果']) and '审批结果' in leave_record else ''
                    days = leave_record['休假天数']
                    
                    # 构建详细记录，只包含必要信息
                    detail_parts = [leave_type]
                    if start_time and start_time != 'nan':
                        detail_parts.append(f"开始:{start_time}")
                    if end_time and end_time != 'nan':
                        detail_parts.append(f"结束:{end_time}")
                    if duration and duration != 'nan':
                        detail_parts.append(f"时长:{duration}")
                    
                    # 将每条记录作为单独的行
                    leave_details.append(" ".join(detail_parts))
                    total_days += days
                    
                    # 检查是否有影响全勤的休假类型
                    if '事假' in leave_type or '病假' in leave_type:
                        has_unpaid_leave = True
                
                # 根据休假类型更新考勤情况
                if has_unpaid_leave:
                    if '考勤情况' in result_df.columns:
                        result_df.at[index, '考勤情况'] = '非全勤'
                    if '全勤' in result_df.columns:
                        result_df.at[index, '全勤'] = 0
                else:
                    if '考勤情况' in result_df.columns:
                        result_df.at[index, '考勤情况'] = '全勤'
                
                # 在备注列中记录详细的休假信息，每条记录分行显示
                if '备注' in result_df.columns:
                    current_note = str(result_df.at[index, '备注']) if pd.notna(result_df.at[index, '备注']) else ''
                    # 使用换行符分隔每条休假记录
                    leave_note = f"休假共{total_days}天:\n" + "\n".join([f"• {detail}" for detail in leave_details])
                    if current_note and current_note != 'nan':
                        result_df.at[index, '备注'] = f"{current_note}\n{leave_note}"
                    else:
                        result_df.at[index, '备注'] = leave_note
        
        # 统计有休假记录的员工数量
        employees_with_leave = leave_data['创建人'].nunique()
        st.success(f"已处理 {employees_with_leave} 名员工的休假数据，更新到现有列中")
    
    return result_df

def process_overtime_data(result_df, overtime_data):
    """处理加班数据并更新到工资表现有列中，根据日期类型填入不同列"""
    if overtime_data is not None:
        # 检查必要的列是否存在
        required_overtime_columns = ['创建人', '时长']
        missing_columns = [col for col in required_overtime_columns if col not in overtime_data.columns]
        
        if missing_columns:
            st.error(f"加班数据文件缺少必要的列: {', '.join(missing_columns)}")
            st.error(f"当前文件包含的列: {', '.join(overtime_data.columns.tolist())}")
            st.error("请确保加班数据文件包含以下列：创建人、时长")
            return result_df
        
        # 显示所有加班记录，不再过滤审批结果
        st.info(f"正在处理 {len(overtime_data)} 条加班记录")
        
        # 处理时长数据，统一转换为小时数
        def parse_overtime_duration(duration):
            if pd.isna(duration):
                return 0
            if isinstance(duration, (int, float)):
                return float(duration)
            duration_str = str(duration).strip()
            if '小时' in duration_str or 'h' in duration_str.lower():
                return float(duration_str.replace('小时', '').replace('h', '').replace('H', ''))
            elif '天' in duration_str:
                days = float(duration_str.replace('天', ''))
                return days * 8  # 按8小时工作日计算
            else:
                try:
                    return float(duration_str)
                except:
                    return 0
        
        overtime_data['加班时间'] = overtime_data['时长'].apply(parse_overtime_duration)
        
        # 按员工姓名分组，收集详细的加班记录
        for index, row in result_df.iterrows():
            employee_name = row['姓名']
            employee_overtime = overtime_data[overtime_data['创建人'] == employee_name]
            
            if not employee_overtime.empty:
                # 分类统计不同类型的加班时间
                weekday_hours = 0  # 平日加班
                weekend_hours = 0  # 休息日加班
                holiday_hours = 0  # 法定节假日加班
                
                # 收集详细的加班记录
                overtime_details = []
                for _, overtime_row in employee_overtime.iterrows():
                    overtime_hours = overtime_row['加班时间']
                    
                    # 尝试解析加班日期
                    overtime_date = None
                    date_type = "工作日"  # 默认为工作日
                    
                    # 从多个可能的日期列中获取日期
                    date_columns = ['开始时间', '日期', '加班日期', '申请日期']
                    for col in date_columns:
                        if col in overtime_row and pd.notna(overtime_row[col]):
                            overtime_date = parse_date_from_string(overtime_row[col])
                            if overtime_date:
                                break
                    
                    # 判断日期类型并分类统计
                    if overtime_date:
                        is_special, date_type = is_holiday_or_weekend(overtime_date)
                        if date_type == "法定节假日":
                            holiday_hours += overtime_hours
                        elif date_type == "休息日":
                            weekend_hours += overtime_hours
                        else:
                            weekday_hours += overtime_hours
                    else:
                        # 如果无法解析日期，默认为平日加班
                        weekday_hours += overtime_hours
                    
                    # 构建详细记录
                    if overtime_date:
                        detail = f"{overtime_date.strftime('%Y-%m-%d')}({date_type}) {overtime_row['时长']}({overtime_hours}小时)"
                    else:
                        detail = f"{overtime_row['时长']}({overtime_hours}小时)"
                    
                    overtime_details.append(detail)
                
                # 更新不同类型的加班时间到对应列
                if weekday_hours > 0 and '平日累计时间' in result_df.columns:
                    current_hours = result_df.at[index, '平日累计时间'] if pd.notna(result_df.at[index, '平日累计时间']) else 0
                    result_df.at[index, '平日累计时间'] = float(current_hours) + weekday_hours
                
                if weekend_hours > 0 and '休息日累计时间' in result_df.columns:
                    current_hours = result_df.at[index, '休息日累计时间'] if pd.notna(result_df.at[index, '休息日累计时间']) else 0
                    result_df.at[index, '休息日累计时间'] = float(current_hours) + weekend_hours
                
                if holiday_hours > 0 and '法定节假日累计时间' in result_df.columns:
                    current_hours = result_df.at[index, '法定节假日累计时间'] if pd.notna(result_df.at[index, '法定节假日累计时间']) else 0
                    result_df.at[index, '法定节假日累计时间'] = float(current_hours) + holiday_hours
                
                # 在备注列中记录详细的加班信息，每条记录分行显示
                if '备注' in result_df.columns:
                    current_note = str(result_df.at[index, '备注']) if pd.notna(result_df.at[index, '备注']) else ''
                    
                    # 构建加班统计信息
                    total_hours = weekday_hours + weekend_hours + holiday_hours
                    overtime_summary = []
                    if weekday_hours > 0:
                        overtime_summary.append(f"平日{weekday_hours}小时")
                    if weekend_hours > 0:
                        overtime_summary.append(f"休息日{weekend_hours}小时")
                    if holiday_hours > 0:
                        overtime_summary.append(f"法定节假日{holiday_hours}小时")
                    
                    summary_text = "、".join(overtime_summary)
                    overtime_note = f"加班共{total_hours}小时({summary_text}):\n" + "\n".join([f"• {detail}" for detail in overtime_details])
                    
                    if current_note and current_note != 'nan':
                        result_df.at[index, '备注'] = f"{current_note}\n{overtime_note}"
                    else:
                        result_df.at[index, '备注'] = overtime_note
        
        # 统计有加班记录的员工数量和日期解析情况
        employees_with_overtime = overtime_data['创建人'].nunique()
        
        # 统计日期解析成功率
        date_parsed_count = 0
        for _, row in overtime_data.iterrows():
            date_columns = ['开始时间', '日期', '加班日期', '申请日期']
            for col in date_columns:
                if col in row and pd.notna(row[col]):
                    if parse_date_from_string(row[col]):
                        date_parsed_count += 1
                        break
        
        st.success(f"已处理 {employees_with_overtime} 名员工的加班数据，按日期类型分类填入对应列")
        if date_parsed_count < len(overtime_data):
            st.warning(f"有 {len(overtime_data) - date_parsed_count} 条记录无法解析日期，已按平日加班处理")
    
    return result_df

def merge_to_salary_sheet(salary_df, leave_df=None, overtime_df=None):
    """将休假和加班数据更新到工资表现有列中，保持原始格式不变"""
    result_df = salary_df.copy()
    
    # 处理休假数据
    if leave_df is not None and not leave_df.empty:
        st.info("正在处理休假数据...")
        result_df = process_leave_data(result_df, leave_df)
    
    # 处理加班数据
    if overtime_df is not None and not overtime_df.empty:
        st.info("正在处理加班数据...")
        result_df = process_overtime_data(result_df, overtime_df)
    
    return result_df

def save_salary_sheet_with_format(result_df, template_path):
    """保存工资表，完整保留模板格式、标题行和公式"""
    try:
        # 加载原始模板工作簿
        wb = load_workbook(template_path)
        ws = wb.active
        
        # 数据从第6行开始（第5行是标题行）
        start_row = 6
        
        # 获取列名映射（第5行是标题行）
        header_row = 5
        col_mapping = {}
        for col_idx, cell in enumerate(ws[header_row], 1):
            if cell.value:
                col_mapping[str(cell.value).strip()] = col_idx
        
        # 清除现有数据行（保留格式和公式）
        max_row = ws.max_row
        for row_idx in range(start_row, max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                # 只清除非公式单元格的值，保留所有公式
                if cell.data_type != 'f':  # 'f' 表示公式类型
                    cell.value = None
        
        # 填入新数据
        for df_row_idx, (_, row_data) in enumerate(result_df.iterrows()):
            excel_row = start_row + df_row_idx
            
            # 为每一列填入数据
            for col_name, value in row_data.items():
                if col_name in col_mapping:
                    col_idx = col_mapping[col_name]
                    cell = ws.cell(row=excel_row, column=col_idx)
                    
                    # 只填入非公式单元格，保护现有公式
                    if cell.data_type != 'f':  # 不覆盖公式单元格
                        # 处理不同类型的值
                        if pd.isna(value) or value == 'nan':
                            cell.value = None
                        elif isinstance(value, str) and value.strip() == '':
                            cell.value = None
                        else:
                            cell.value = value
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
        
    except Exception as e:
        st.error(f"保存工资表时出错: {str(e)}")
        return None

def main():
    st.set_page_config(
        page_title="智能工资表生成系统",
        page_icon="💰",
        layout="wide",
        initial_sidebar_state="expanded"
    )
    
    # 自定义CSS样式
    st.markdown("""
    <style>
    /* 主题色彩定义 */
    :root {
        --primary-color: #1f77b4;
        --secondary-color: #ff7f0e;
        --success-color: #2ca02c;
        --warning-color: #ff9800;
        --error-color: #d62728;
        --background-gradient: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        --card-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
    }
    
    /* 主标题样式 */
    .main-title {
        background: var(--background-gradient);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        background-clip: text;
        font-size: 3rem;
        font-weight: 700;
        text-align: center;
        margin-bottom: 2rem;
        text-shadow: 2px 2px 4px rgba(0,0,0,0.1);
    }
    
    /* 卡片样式 */
    .custom-card {
        background: white;
        padding: 1.5rem;
        border-radius: 15px;
        box-shadow: var(--card-shadow);
        border: 1px solid #e0e0e0;
        margin-bottom: 1rem;
        transition: transform 0.2s ease, box-shadow 0.2s ease;
    }
    
    .custom-card:hover {
        transform: translateY(-2px);
        box-shadow: 0 8px 15px rgba(0, 0, 0, 0.15);
    }
    
    /* 状态指示器 */
    .status-indicator {
        display: inline-flex;
        align-items: center;
        padding: 0.5rem 1rem;
        border-radius: 25px;
        font-weight: 600;
        font-size: 0.9rem;
        margin: 0.25rem;
    }
    
    .status-success {
        background: linear-gradient(135deg, #4CAF50, #45a049);
        color: white;
    }
    
    .status-warning {
        background: linear-gradient(135deg, #FF9800, #f57c00);
        color: white;
    }
    
    .status-error {
        background: linear-gradient(135deg, #f44336, #d32f2f);
        color: white;
    }
    
    /* 统计卡片 */
    .metric-card {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        padding: 1.5rem;
        border-radius: 15px;
        text-align: center;
        box-shadow: var(--card-shadow);
        margin-bottom: 1rem;
    }
    
    .metric-number {
        font-size: 2.5rem;
        font-weight: 700;
        margin-bottom: 0.5rem;
    }
    
    .metric-label {
        font-size: 1rem;
        opacity: 0.9;
    }
    
    /* 按钮样式 */
    .stButton > button {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        border: none;
        border-radius: 25px;
        padding: 0.75rem 2rem;
        font-weight: 600;
        font-size: 1.1rem;
        transition: all 0.3s ease;
        box-shadow: 0 4px 15px rgba(102, 126, 234, 0.4);
    }
    
    .stButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 6px 20px rgba(102, 126, 234, 0.6);
    }
    
    /* 文件上传区域 */
    .uploadedFile {
        border: 2px dashed #667eea;
        border-radius: 15px;
        padding: 1rem;
        text-align: center;
        background: linear-gradient(135deg, rgba(102, 126, 234, 0.1), rgba(118, 75, 162, 0.1));
    }
    
    /* 侧边栏样式 */
    .css-1d391kg {
        background: linear-gradient(180deg, #f8f9fa 0%, #e9ecef 100%);
    }
    
    /* 数据表格样式 */
    .stDataFrame {
        border-radius: 10px;
        overflow: hidden;
        box-shadow: var(--card-shadow);
    }
    
    /* 进度条样式 */
    .stProgress > div > div {
        background: linear-gradient(90deg, #667eea 0%, #764ba2 100%);
    }
    
    /* 展开器样式 */
    .streamlit-expanderHeader {
        background: linear-gradient(135deg, rgba(102, 126, 234, 0.1), rgba(118, 75, 162, 0.1));
        border-radius: 10px;
        font-weight: 600;
    }
    
    /* 分隔线样式 */
    .custom-divider {
        height: 3px;
        background: linear-gradient(90deg, #667eea 0%, #764ba2 100%);
        border: none;
        border-radius: 2px;
        margin: 2rem 0;
    }
    
    /* 功能介绍卡片 */
    .feature-card {
        background: white;
        padding: 1.5rem;
        border-radius: 12px;
        text-align: center;
        box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        border: 1px solid #f0f0f0;
        transition: all 0.3s ease;
        height: 100%;
        margin-bottom: 1rem;
    }
    
    .feature-card:hover {
        transform: translateY(-5px);
        box-shadow: 0 8px 25px rgba(102, 126, 234, 0.15);
        border-color: #667eea;
    }
    
    .feature-icon {
        font-size: 2.5rem;
        margin-bottom: 1rem;
        display: block;
    }
    
    .feature-card h4 {
        color: #333;
        margin-bottom: 0.5rem;
        font-weight: 600;
    }
    
    .feature-card p {
        color: #666;
        font-size: 0.9rem;
        line-height: 1.4;
        margin: 0;
    }
    
    .feature-title {
        color: #667eea;
        font-weight: 600;
        font-size: 1.1rem;
        margin-bottom: 0.5rem;
    }
    
    /* 响应式设计 */
    @media (max-width: 768px) {
        .main-title {
            font-size: 2rem;
        }
        .metric-number {
            font-size: 2rem;
        }
    }
    </style>
    """, unsafe_allow_html=True)
    
    # 主标题
    st.markdown('<h1 class="main-title">💰 智能工资表生成系统</h1>', unsafe_allow_html=True)
    st.markdown('<hr class="custom-divider">', unsafe_allow_html=True)
    
    # 侧边栏
    with st.sidebar:
        st.markdown("### 🎛️ 控制面板")
        st.markdown("---")
        
        # 加载工资表模板
        st.markdown("#### 📊 工资表模板")
        salary_template, template_path = load_salary_template()
        
        if salary_template is not None:
            st.markdown('<div class="status-indicator status-success">✅ 模板加载成功</div>', unsafe_allow_html=True)
            st.markdown(f"**员工数量:** {len(salary_template)} 人")
        else:
            st.markdown('<div class="status-indicator status-error">❌ 模板加载失败</div>', unsafe_allow_html=True)
            st.stop()
        
        st.markdown("---")
        
        # 文件上传区域
        st.markdown("#### 📁 数据文件上传")
        
        # 休假数据上传
        st.markdown("**🏖️ 休假表**")
        leave_file = st.file_uploader(
            "上传休假表",
            type=['xlsx', 'xls'],
            key="leave_file",
            help="可选上传，包含员工休假信息的Excel文件"
        )
        
        # 加班数据上传
        st.markdown("**⏰ 加班表**")
        overtime_file = st.file_uploader(
            "上传加班表",
            type=['xlsx', 'xls'],
            key="overtime_file",
            help="可选上传，包含员工加班信息的Excel文件"
        )
    
    # 主内容区域
    # 系统功能简介
    st.markdown("""
    <div class="custom-card" style="text-align: center; background: linear-gradient(135deg, rgba(102, 126, 234, 0.05), rgba(118, 75, 162, 0.05));">
        <h4 style="color: #667eea; margin-bottom: 1rem;">💡 系统功能</h4>
        <p style="margin-bottom: 0;">智能考勤判断 • 节假日识别 • 公式保护 • 详细备注</p>
    </div>
    """, unsafe_allow_html=True)
    
    # 使用说明
    with st.expander("💡 使用说明", expanded=False):
        st.markdown("""
        **操作步骤：**
        1. 系统自动加载工资表模板
        2. 可选上传请假表和加班表
        3. 点击生成工资表按钮
        4. 下载生成的工资表文件
        
        **数据格式：**
        - 请假表：姓名、请假类型、时长
        - 加班表：姓名、加班日期、时长
        """)
    
    st.markdown('<hr class="custom-divider">', unsafe_allow_html=True)
    
    # 数据状态
    st.markdown("### 📊 数据状态")
    
    # 简化的统计信息
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.metric("员工总数", len(salary_template))
    
    with col2:
        leave_count = len(load_leave_data(leave_file)) if leave_file else 0
        st.metric("请假记录", leave_count)
    
    with col3:
        overtime_count = len(load_overtime_data(overtime_file)) if overtime_file else 0
        st.metric("加班记录", overtime_count)
    
    # 数据预览（简化）
    if leave_file or overtime_file:
        with st.expander("📋 数据预览", expanded=False):
            if leave_file:
                st.write("**请假数据：**")
                leave_preview = load_leave_data(leave_file)
                if leave_preview is not None and not leave_preview.empty:
                    st.dataframe(leave_preview.head(3), use_container_width=True)
            
            if overtime_file:
                st.write("**加班数据：**")
                overtime_preview = load_overtime_data(overtime_file)
                if overtime_preview is not None and not overtime_preview.empty:
                    st.dataframe(overtime_preview.head(3), use_container_width=True)
    
    # 生成工资表按钮
    st.markdown('<hr class="custom-divider">', unsafe_allow_html=True)
    
    # 生成按钮区域
    st.markdown("### 🚀 生成工资表")
    
    # 检查是否可以生成
    can_generate = salary_template is not None
    
    if can_generate:
        st.markdown("""
        <div class="custom-card" style="text-align: center; background: linear-gradient(135deg, rgba(102, 126, 234, 0.05), rgba(118, 75, 162, 0.05));">
            <h4 style="color: #667eea; margin-bottom: 1rem;">🎯 准备就绪</h4>
            <p style="margin-bottom: 1.5rem;">系统已准备好生成工资表，点击下方按钮开始处理</p>
        </div>
        """, unsafe_allow_html=True)
    else:
        st.markdown("""
        <div class="custom-card" style="text-align: center; background: linear-gradient(135deg, rgba(255, 152, 0, 0.05), rgba(255, 87, 34, 0.05));">
            <h4 style="color: #ff9800; margin-bottom: 1rem;">⚠️ 请检查配置</h4>
            <p style="margin-bottom: 1.5rem;">请确保工资表模板已正确加载</p>
        </div>
        """, unsafe_allow_html=True)
    
    col1, col2, col3 = st.columns([1, 2, 1])
    
    with col2:
        if st.button("🚀 开始生成工资表", type="primary", use_container_width=True, disabled=not can_generate):
            # 创建进度条
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            try:
                # 步骤1: 加载数据
                status_text.text("📂 正在加载数据文件...")
                progress_bar.progress(20)
                
                leave_data = load_leave_data(leave_file) if leave_file else None
                overtime_data = load_overtime_data(overtime_file) if overtime_file else None
                
                # 步骤2: 处理数据
                status_text.text("⚙️ 正在处理员工数据...")
                progress_bar.progress(50)
                
                final_salary_sheet = merge_to_salary_sheet(
                    salary_template, 
                    leave_data, 
                    overtime_data
                )
                
                # 步骤3: 生成Excel文件
                status_text.text("📊 正在生成Excel文件...")
                progress_bar.progress(80)
                
                excel_data = save_salary_sheet_with_format(final_salary_sheet, template_path)
                
                if excel_data is None:
                    st.error("❌ 生成Excel文件失败，请检查模板格式")
                    st.stop()
                
                # 步骤4: 完成
                status_text.text("✅ 工资表生成完成！")
                progress_bar.progress(100)
                
                # 成功提示
                st.balloons()
                st.markdown("""
                <div class="custom-card" style="text-align: center; background: linear-gradient(135deg, rgba(76, 175, 80, 0.1), rgba(69, 160, 73, 0.1)); border: 2px solid #4CAF50;">
                    <h3 style="color: #4CAF50; margin-bottom: 1rem;">🎉 生成成功！</h3>
                    <p style="margin-bottom: 1rem;">工资表已成功生成，包含所有员工的考勤和加班信息</p>
                </div>
                """, unsafe_allow_html=True)
                
                # 显示最终工资表
                st.markdown("### 📋 最终工资表预览")
                
                with st.expander("📊 查看完整工资表", expanded=True):
                    st.dataframe(final_salary_sheet, use_container_width=True, height=500)
                
                # 下载按钮
                st.markdown("### 📥 下载文件")
                
                col_download1, col_download2, col_download3 = st.columns([1, 2, 1])
                with col_download2:
                    st.download_button(
                        label="📥 下载完整工资表",
                        data=excel_data,
                        file_name=f"工资表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                
            except Exception as e:
                st.error(f"❌ 生成过程中出现错误: {str(e)}")
                progress_bar.empty()
                status_text.empty()
    
    # 页脚信息
    st.markdown('<hr class="custom-divider">', unsafe_allow_html=True)
    
    # 数据格式说明
    with st.expander("📝 数据格式说明", expanded=False):
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("""
            <div class="custom-card">
                <h4 style="color: #667eea; margin-bottom: 1rem;">📋 请假数据格式</h4>
                <ul style="margin-left: 1rem;">
                    <li><strong>姓名：</strong>员工姓名</li>
                    <li><strong>请假类型：</strong>事假、病假、年假等</li>
                    <li><strong>请假时长：</strong>支持多种格式</li>
                    <li style="margin-left: 1rem; color: #666;">• 1天、8小时、1.5天</li>
                    <li style="margin-left: 1rem; color: #666;">• 0.5天、4小时等</li>
                </ul>
            </div>
            """, unsafe_allow_html=True)
        
        with col2:
            st.markdown("""
            <div class="custom-card">
                <h4 style="color: #667eea; margin-bottom: 1rem;">⏰ 加班数据格式</h4>
                <ul style="margin-left: 1rem;">
                    <li><strong>姓名：</strong>员工姓名</li>
                    <li><strong>加班日期：</strong>多种日期格式</li>
                    <li style="margin-left: 1rem; color: #666;">• YYYY-MM-DD</li>
                    <li style="margin-left: 1rem; color: #666;">• MM/DD/YYYY等</li>
                    <li><strong>加班时长：</strong>2小时、1.5小时等</li>
                </ul>
            </div>
            """, unsafe_allow_html=True)
        
        st.markdown("""
        <div class="custom-card" style="background: linear-gradient(135deg, rgba(255, 193, 7, 0.1), rgba(255, 152, 0, 0.1)); border-left: 4px solid #ffc107;">
            <h4 style="color: #ff9800; margin-bottom: 1rem;">⚠️ 重要提示</h4>
            <ul style="margin-left: 1rem;">
                <li>系统会自动识别法定节假日和休息日</li>
                <li>请假会影响考勤情况和全勤工资计算</li>
                <li>所有Excel公式会自动保留在生成的文件中</li>
                <li>支持批量处理多个员工的考勤数据</li>
            </ul>
        </div>
        """, unsafe_allow_html=True)
    

    
    # 页脚
    st.markdown('<hr class="custom-divider">', unsafe_allow_html=True)
    st.markdown("""
    <div style="text-align: center; padding: 1rem; color: #666;">
        <p>💼 智能工资表生成系统 v2.0 | 智能考勤管理</p>
    </div>
    """, unsafe_allow_html=True)

if __name__ == "__main__":
    main()